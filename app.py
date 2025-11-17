import os
from dotenv import load_dotenv
import zipfile
import tempfile
import re

from flask import Flask, render_template, request, send_file
from openpyxl import load_workbook
import cloudinary
import cloudinary.uploader

# Load environment variables from .env file
load_dotenv()

app = Flask(__name__)

# -------------------------
# Cloudinary Configuration
# -------------------------
cloudinary.config(
    cloud_name=os.environ.get("CLOUDINARY_CLOUD_NAME"),
    api_key=os.environ.get("CLOUDINARY_API_KEY"),
    api_secret=os.environ.get("CLOUDINARY_API_SECRET"),
    secure=True
)

# -------------------------
# Extract English Text
# -------------------------
def extract_english(text):
    if text is None:
        return ""
    base = os.path.splitext(str(text))[0]
    eng = re.sub(r'[^A-Za-z0-9 ]+', '', base)
    return eng.strip().lower()

# -------------------------
# Recursively find all images
# -------------------------
def find_images_recursively(root):
    img_paths = {}
    print("\n========= ZIP SCAN START =========")
    for dirpath, _, filenames in os.walk(root):
        for filename in filenames:

            # 🔥 skip macOS crap files
            if filename.startswith("._"):
                print(f"Skipping macOS hidden file: {filename}")
                continue

            if filename.lower().endswith((".jpg", ".jpeg", ".png")):
                eng = extract_english(filename)
                full_path = os.path.join(dirpath, filename)
                img_paths[eng] = full_path
                print(f"Found image: '{filename}' → key = '{eng}'")
    print("========= ZIP SCAN END =========\n")
    return img_paths

# -------------------------
# Main Processor
# -------------------------
def process_excel_and_zip(excel_file, zip_file):

    # Save excel
    temp_excel = tempfile.mktemp(suffix=".xlsx")
    excel_file.save(temp_excel)

    # Extract zip
    temp_dir = tempfile.mkdtemp()
    with zipfile.ZipFile(zip_file, 'r') as z:
        z.extractall(temp_dir)

    # Scan images
    image_map = find_images_recursively(temp_dir)

    # Load excel
    wb = load_workbook(temp_excel)
    ws = wb.active

    print("========= MATCHING START =========")
    for row in ws.iter_rows(min_row=2):
        attraction_name = str(row[1].value)
        eng = extract_english(attraction_name)

        print(f"\nExcel attraction: '{attraction_name}' → key = '{eng}'")

        if eng in image_map:
            print(f"✔ MATCHED with image key = '{eng}'")
            local_img_path = image_map[eng]

            uploaded = cloudinary.uploader.upload(
                local_img_path,
                folder="pandahoho/covers",
                public_id=eng,
                overwrite=True
            )

            row[3].value = uploaded["secure_url"]
        else:
            print(f"❌ NO MATCH for key: '{eng}'")
            row[3].value = ""

    print("========= MATCHING END =========\n")

    output_path = tempfile.mktemp(suffix=".xlsx")
    wb.save(output_path)
    return output_path

# -------------------------
# Routes
# -------------------------
@app.route("/")
def index():
    return render_template("index.html")

@app.route("/upload", methods=["POST"])
def upload():
    if "excel_file" not in request.files or "zip_file" not in request.files:
        return "❌ Missing Excel or ZIP file", 400

    excel_file = request.files["excel_file"]
    zip_file = request.files["zip_file"]

    try:
        output_path = process_excel_and_zip(excel_file, zip_file)
    except Exception as e:
        print("[ERROR]", e)
        return f"❌ Error: {str(e)}", 500

    return send_file(
        output_path,
        as_attachment=True,
        download_name="updated.xlsx"
    )

if __name__ == "__main__":
    # Use PORT from environment variable (Railway provides this)
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)